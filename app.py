from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import json
import os
from datetime import datetime
import tempfile
from werkzeug.utils import secure_filename
import sys
import traceback

# Add the current directory to Python path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from test_case_generator import get_ai_provider, AI_CONFIG, generate_test_cases, parse_markdown_table, save_to_excel

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure templates and static directories exist
os.makedirs('templates', exist_ok=True)
os.makedirs('static', exist_ok=True)
os.makedirs('static/css', exist_ok=True)
os.makedirs('static/js', exist_ok=True)

def make_json_serializable(obj):
    """Convert objects to JSON serializable format"""
    if obj is None:
        return None
    elif isinstance(obj, pd.DataFrame):
        try:
            return obj.to_dict('records')
        except:
            return []
    elif isinstance(obj, (pd.Series, pd.Index)):
        try:
            return obj.tolist()
        except:
            return []
    elif isinstance(obj, (datetime, pd.Timestamp)):
        return obj.isoformat()
    elif isinstance(obj, (int, float, str, bool)):
        return obj
    elif isinstance(obj, (list, tuple)):
        return [make_json_serializable(item) for item in obj]
    elif isinstance(obj, dict):
        return {key: make_json_serializable(value) for key, value in obj.items()}
    else:
        try:
            return str(obj)
        except:
            return None

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/api/test')
def api_test():
    """Test endpoint to check AI provider setup"""
    try:
        print("Testing AI provider setup...")
        
        # Test AI_CONFIG
        print(f"AI_CONFIG keys: {list(AI_CONFIG.keys())}")
        print(f"Current provider: {AI_CONFIG.get('provider', 'Not set')}")
        
        # Test specific provider
        provider_name = 'gemini'
        print(f"Testing provider: {provider_name}")
        
        try:
            provider = get_ai_provider(provider_name)
            print(f"Provider initialized: {type(provider)}")
            
            # Test a simple generation
            test_story = "As a user, I want to login so that I can access my account."
            print("Testing simple story generation...")
            result = provider.generate_test_cases(test_story, "TEST001")
            print(f"Generation result length: {len(result) if result else 0}")
            
            return jsonify({
                'success': True,
                'message': 'AI provider test successful',
                'provider': provider_name,
                'test_result_length': len(result) if result else 0
            })
            
        except Exception as e:
            print(f"Provider test failed: {e}")
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': f'Provider test failed: {str(e)}',
                'provider': provider_name
            }), 500
            
    except Exception as e:
        print(f"Test endpoint error: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate', methods=['POST'])
def api_generate():
    """API endpoint for generating test cases"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        story_id = data.get('story_id', 'US001')
        story_title = data.get('story_title', 'Custom User Story')
        user_story = data.get('user_story', '')
        acceptance_criteria = data.get('acceptance_criteria', '')
        ai_provider = data.get('ai_provider', 'gemini')
        
        if not user_story.strip():
            return jsonify({'error': 'User story is required'}), 400
        
        # Combine user story and acceptance criteria
        full_story = f"{user_story}\n\nAcceptance Criteria:\n{acceptance_criteria}"
        
        print(f"Debug: Using AI provider: {ai_provider}")
        print(f"Debug: Full story length: {len(full_story)}")
        
        # Get AI provider
        try:
            provider = get_ai_provider(ai_provider)
            if not provider:
                return jsonify({'error': f'AI provider {ai_provider} not available'}), 400
        except Exception as e:
            print(f"Debug: Error getting AI provider: {e}")
            traceback.print_exc()
            return jsonify({'error': f'Failed to initialize AI provider: {str(e)}'}), 500
        
        # Generate test cases
        try:
            print("Debug: Generating test cases...")
            test_cases = provider.generate_test_cases(full_story, story_id)
            print(f"Debug: Test cases generated, length: {len(test_cases) if test_cases else 0}")
            
            if not test_cases:
                return jsonify({'error': 'Failed to generate test cases'}), 500
        except Exception as e:
            print(f"Debug: Error generating test cases: {e}")
            traceback.print_exc()
            return jsonify({'error': f'Failed to generate test cases: {str(e)}'}), 500
        
        # Parse markdown table if present
        try:
            parsed_cases = parse_markdown_table(test_cases)
            print(f"Debug: Markdown parsed successfully")
            print(f"Debug: Parsed cases type: {type(parsed_cases)}")
            if parsed_cases is not None:
                print(f"Debug: Parsed cases shape: {parsed_cases.shape if hasattr(parsed_cases, 'shape') else 'No shape'}")
            
            # Convert DataFrame to JSON-serializable format if it exists
            if parsed_cases is not None:
                try:
                    # Convert DataFrame to records (list of dictionaries)
                    parsed_cases = parsed_cases.to_dict('records')
                    print(f"Debug: Converted DataFrame to {len(parsed_cases)} records")
                except Exception as convert_error:
                    print(f"Debug: Error converting DataFrame: {convert_error}")
                    parsed_cases = None
            else:
                print("Debug: No parsed cases to convert")
                
        except Exception as e:
            print(f"Debug: Error parsing markdown: {e}")
            traceback.print_exc()
            parsed_cases = None
        
        # Ensure all data is JSON serializable
        try:
            # Use utility function to make data JSON serializable
            safe_parsed_cases = make_json_serializable(parsed_cases)
            
            # Test JSON serialization
            test_response = {
                'success': True,
                'test_cases': test_cases,
                'parsed_cases': safe_parsed_cases if safe_parsed_cases is not None else [],
                'story_id': story_id,
                'story_title': story_title,
                'timestamp': datetime.now().isoformat()
            }
            
            # Test if we can serialize this
            json.dumps(test_response)
            print("Debug: JSON serialization test passed")
            
            return jsonify(test_response)
            
        except Exception as json_error:
            print(f"Debug: JSON serialization failed: {json_error}")
            # Fallback response without problematic data
            return jsonify({
                'success': True,
                'test_cases': test_cases,
                'parsed_cases': [],
                'story_id': story_id,
                'story_title': story_title,
                'timestamp': datetime.now().isoformat(),
                'note': 'Parsed cases could not be serialized'
            })
        
    except Exception as e:
        print(f"Debug: Unexpected error in api_generate: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/export', methods=['POST'])
def api_export():
    """API endpoint for exporting test cases to Excel"""
    temp_file_path = None
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        test_cases = data.get('test_cases', '')
        story_id = data.get('story_id', 'US001')
        story_title = data.get('story_title', 'Custom User Story')
        
        if not test_cases.strip():
            return jsonify({'error': 'Test cases are required'}), 400
        
        print(f"Debug: Exporting test cases for {story_id}: {story_title}")
        print(f"Debug: Test cases length: {len(test_cases)}")
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            temp_file_path = tmp_file.name
            
            try:
                # Parse the test cases to get a DataFrame
                parsed_df = parse_markdown_table(test_cases)
                
                if parsed_df is not None:
                    # Format data as expected by save_to_excel
                    export_data = [{
                        "story_id": story_id,
                        "story_title": story_title,
                        "test_cases": parsed_df
                    }]
                    
                    print(f"Debug: Exporting {len(export_data)} stories with {len(parsed_df)} test cases")
                    save_to_excel(export_data, tmp_file.name)
                else:
                    # Fallback: create a simple Excel file with raw text
                    print("Debug: No parsed data, creating simple Excel file")
                    create_simple_excel(test_cases, story_id, story_title, tmp_file.name)
                
                # Return file
                return send_file(
                    tmp_file.name,
                    as_attachment=True,
                    download_name=f'test_cases_{story_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
            except Exception as excel_error:
                print(f"Debug: Excel creation error: {excel_error}")
                traceback.print_exc()
                
                # Try to create a simple text-based Excel file as fallback
                try:
                    create_simple_excel(test_cases, story_id, story_title, tmp_file.name)
                    return send_file(
                        tmp_file.name,
                        as_attachment=True,
                        download_name=f'test_cases_{story_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                except Exception as fallback_error:
                    print(f"Debug: Fallback Excel creation also failed: {fallback_error}")
                    raise fallback_error
            
    except Exception as e:
        print(f"Debug: Error in api_export: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    
    finally:
        # Clean up temporary file if it exists and there was an error
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                print(f"Debug: Cleaned up temporary file: {temp_file_path}")
            except Exception as cleanup_error:
                print(f"Debug: Failed to cleanup temporary file: {cleanup_error}")

def create_simple_excel(test_cases, story_id, story_title, filename):
    """Create a simple Excel file with raw test case data"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{story_id[:30]}"
    
    # Add headers
    headers = ["Story ID", "Story Title", "Test Cases"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    ws.cell(row=2, column=1, value=story_id)
    ws.cell(row=2, column=2, value=story_title)
    ws.cell(row=2, column=3, value=test_cases)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    print(f"Debug: Simple Excel file created: {filename}")

@app.route('/api/providers')
def api_providers():
    """API endpoint for getting available AI providers"""
    try:
        providers = list(AI_CONFIG.keys())
        current = AI_CONFIG.get('provider', 'gemini')
        print(f"Debug: Available providers: {providers}, Current: {current}")
        return jsonify({
            'providers': providers,
            'current': current
        })
    except Exception as e:
        print(f"Debug: Error in api_providers: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/examples')
def api_examples():
    """API endpoint for getting example user stories"""
    try:
        from test_case_generator import EXAMPLE_STORIES
        print(f"Debug: Loaded {len(EXAMPLE_STORIES)} example stories")
        return jsonify({'examples': EXAMPLE_STORIES})
    except Exception as e:
        print(f"Debug: Error in api_examples: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 