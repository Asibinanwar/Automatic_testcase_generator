import pandas as pd
import openai
import re
import json
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import argparse
from datetime import datetime
import time
from abc import ABC, abstractmethod

# AI Provider Configuration
AI_CONFIG = {
    "provider": "gemini",  # Changed to use Gemini as default
    "openai": {
        "api_key": "aH9NYzXLrTgUEmxBhoFGvduqBwL1iTzzgfc3Iw1Xz9t45JhrSZrnJQQJ99BGACYeBjFXJ3w3AAABACOGSkVF",
        "model": "gpt-4.1"
    },
    "anthropic": {
        "api_key": "your-anthropic-api-key-here",
        "model": "claude-3-sonnet-20240229"
    },
    "gemini": {
        "api_key": "AIzaSyCG9bq2U8RkpzJru6lDMZ-6kp0NWV22xKs",
        "model": "gemini-2.0-flash"  # Faster model with higher limits
    },
    "azure_openai": {
        "api_key": "your-azure-openai-api-key-here",
        "endpoint": "https://your-resource.openai.azure.com/",
        "deployment_name": "your-deployment-name",
        "api_version": "2024-02-15-preview"
    }
}

# Jira Configuration (optional)
JIRA_CONFIG = {
    "base_url": "https://your-domain.atlassian.net",
    "email": "your-email@domain.com",
    "api_token": "your-jira-api-token",
    "project_key": "PROJ"  # Your Jira project key
}

# Example user stories for testing
EXAMPLE_STORIES = [
    {
        "id": "US001",
        "title": "Password Reset Functionality",
        "story": """
User Story: As a customer, I want to be able to reset my password so that I can regain access to my account if I forget my credentials.

Acceptance Criteria:
1. User can click "Forgot Password" link on login page
2. User enters their email address in the password reset form
3. System validates email format and existence in database
4. System sends password reset email with secure token
5. User clicks link in email to access password reset page
6. User enters new password and confirms it
7. System validates password strength requirements
8. System updates password in database and invalidates old token
9. User receives confirmation email
10. User can login with new password
"""
    },
    {
        "id": "US002", 
        "title": "User Registration",
        "story": """
User Story: As a new user, I want to register for an account so that I can access the application.

Acceptance Criteria:
1. User can access registration page
2. User can enter required information (name, email, password)
3. System validates email format and uniqueness
4. System validates password strength requirements
5. User receives email verification link
6. User can verify email by clicking link
7. Account is activated after email verification
8. User can login with verified account
"""
    }
]

class AIProvider(ABC):
    """
    Abstract base class for AI providers.
    """
    @abstractmethod
    def generate_test_cases(self, user_story, story_id="TC"):
        """
        Generate test cases for a given user story.
        """
        pass

class OpenAIProvider(AIProvider):
    """
    Concrete implementation for OpenAI API.
    """
    def __init__(self, api_key: str, model: str):
        openai.api_key = api_key
        self.model = model

    def generate_test_cases(self, user_story, story_id="TC"):
        """
        Generate test cases for a given user story.
        """
        prompt = f"""
    Analyze the following user story and acceptance criteria to generate comprehensive test cases.
    
    User Story and Acceptance Criteria:
    {user_story}
    
    Please identify functional areas and generate risk-based test cases. For each test case:
    1. Assign a Risk Level (High/Medium/Low) based on:
       - Complexity of the functionality
       - Business impact if it fails
       - Likelihood of defects based on common patterns
    
    2. Generate more detailed test cases for high-risk areas and fewer for low-risk areas.
    
    3. Output the results as a clean markdown table with these columns:
       | Test Case ID | Area/Feature | Description | Steps | Expected Result | Risk Level | Priority |
       
    4. Use realistic test case IDs (e.g., {story_id}001, {story_id}002, etc.)
    5. Make descriptions clear and actionable
    6. Include both positive and negative test scenarios
    7. Prioritize based on risk level (High=1, Medium=2, Low=3)
    
    Focus on edge cases, error conditions, and integration points for high-risk areas.
    """
        
        try:
            response = openai.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "You are a senior QA engineer specializing in risk-based testing. Generate comprehensive test cases with clear risk assessments."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=2000,
                temperature=0.3
            )
            
            return response.choices[0].message.content.strip()
            
        except Exception as e:
            print(f"Error calling OpenAI API: {e}")
            return None

class AnthropicProvider(AIProvider):
    """
    Concrete implementation for Anthropic API.
    """
    def __init__(self, api_key: str, model: str):
        import anthropic
        self.client = anthropic.Anthropic(api_key=api_key)
        self.model = model

    def generate_test_cases(self, user_story, story_id="TC"):
        """
        Generate test cases for a given user story.
        """
        prompt = f"""
    Analyze the following user story and acceptance criteria to generate comprehensive test cases.
    
    User Story and Acceptance Criteria:
    {user_story}
    
    Please identify functional areas and generate risk-based test cases. For each test case:
    1. Assign a Risk Level (High/Medium/Low) based on:
       - Complexity of the functionality
       - Business impact if it fails
       - Likelihood of defects based on common patterns
    
    2. Generate more detailed test cases for high-risk areas and fewer for low-risk areas.
    
    3. Output the results as a clean markdown table with these columns:
       | Test Case ID | Area/Feature | Description | Steps | Expected Result | Risk Level | Priority |
       
    4. Use realistic test case IDs (e.g., {story_id}001, {story_id}002, etc.)
    5. Make descriptions clear and actionable
    6. Include both positive and negative test scenarios
    7. Prioritize based on risk level (High=1, Medium=2, Low=3)
    
    Focus on edge cases, error conditions, and integration points for high-risk areas.
    """
        
        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=2000,
                temperature=0.3,
                messages=[
                    {"role": "user", "content": prompt}
                ]
            )
            
            return response.content[0].text.strip()
            
        except Exception as e:
            print(f"Error calling Anthropic API: {e}")
            return None

class GeminiProvider(AIProvider):
    """
    Concrete implementation for Gemini API.
    """
    def __init__(self, api_key: str, model: str):
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel(model)

    def generate_test_cases(self, user_story, story_id="TC"):
        """
        Generate test cases for a given user story.
        """
        prompt = f"""
    Analyze the following user story and acceptance criteria to generate comprehensive test cases.
    
    User Story and Acceptance Criteria:
    {user_story}
    
    Please identify functional areas and generate risk-based test cases. For each test case:
    1. Assign a Risk Level (High/Medium/Low) based on:
       - Complexity of the functionality
       - Business impact if it fails
       - Likelihood of defects based on common patterns
    
    2. Generate more detailed test cases for high-risk areas and fewer for low-risk areas.
    
    3. Output the results as a clean markdown table with these columns:
       | Test Case ID | Area/Feature | Description | Steps | Expected Result | Risk Level | Priority |
       
    4. Use realistic test case IDs (e.g., {story_id}001, {story_id}002, etc.)
    5. Make descriptions clear and actionable
    6. Include both positive and negative test scenarios
    7. Prioritize based on risk level (High=1, Medium=2, Low=3)
    
    Focus on edge cases, error conditions, and integration points for high-risk areas.
    """
        
        try:
            response = self.model.generate_content(prompt)
            
            return response.text.strip()
            
        except Exception as e:
            if "429" in str(e) or "quota" in str(e).lower():
                print(f"⚠️  Rate limit hit. Waiting 60 seconds before retry...")
                import time
                time.sleep(60)
                try:
                    response = self.model.generate_content(prompt)
                    return response.text.strip()
                except Exception as retry_e:
                    print(f"Error calling Gemini API after retry: {retry_e}")
                    return None
            else:
                print(f"Error calling Gemini API: {e}")
                return None

class AzureOpenAIProvider(AIProvider):
    """
    Concrete implementation for Azure OpenAI API.
    """
    def __init__(self, api_key: str, endpoint: str, deployment_name: str, api_version: str):
        from openai import AzureOpenAI
        self.client = AzureOpenAI(
            api_key=api_key,
            azure_endpoint=endpoint,
            api_version=api_version
        )
        self.deployment_name = deployment_name

    def generate_test_cases(self, user_story, story_id="TC"):
        """
        Generate test cases for a given user story.
        """
        prompt = f"""
    Analyze the following user story and acceptance criteria to generate comprehensive test cases.
    
    User Story and Acceptance Criteria:
    {user_story}
    
    Please identify functional areas and generate risk-based test cases. For each test case:
    1. Assign a Risk Level (High/Medium/Low) based on:
       - Complexity of the functionality
       - Business impact if it fails
       - Likelihood of defects based on common patterns
    
    2. Generate more detailed test cases for high-risk areas and fewer for low-risk areas.
    
    3. Output the results as a clean markdown table with these columns:
       | Test Case ID | Area/Feature | Description | Steps | Expected Result | Risk Level | Priority |
       
    4. Use realistic test case IDs (e.g., {story_id}001, {story_id}002, etc.)
    5. Make descriptions clear and actionable
    6. Include both positive and negative test scenarios
    7. Prioritize based on risk level (High=1, Medium=2, Low=3)
    
    Focus on edge cases, error conditions, and integration points for high-risk areas.
    """
        
        try:
            response = self.client.chat.completions.create(
                model=self.deployment_name,
                messages=[
                    {"role": "system", "content": "You are a senior QA engineer specializing in risk-based testing. Generate comprehensive test cases with clear risk assessments."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=2000,
                temperature=0.3
            )
            
            return response.choices[0].message.content.strip()
            
        except Exception as e:
            print(f"Error calling Azure OpenAI API: {e}")
            return None

def get_ai_provider(provider_name: str) -> AIProvider:
    """
    Factory method to get an AI provider based on the configuration.
    """
    if provider_name == "openai":
        return OpenAIProvider(AI_CONFIG["openai"]["api_key"], AI_CONFIG["openai"]["model"])
    elif provider_name == "anthropic":
        return AnthropicProvider(AI_CONFIG["anthropic"]["api_key"], AI_CONFIG["anthropic"]["model"])
    elif provider_name == "gemini":
        return GeminiProvider(AI_CONFIG["gemini"]["api_key"], AI_CONFIG["gemini"]["model"])
    elif provider_name == "azure_openai":
        return AzureOpenAIProvider(AI_CONFIG["azure_openai"]["api_key"], AI_CONFIG["azure_openai"]["endpoint"], AI_CONFIG["azure_openai"]["deployment_name"], AI_CONFIG["azure_openai"]["api_version"])
    else:
        raise ValueError(f"Provider '{provider_name}' not found in AI_CONFIG.")

def get_jira_stories(jql_query=None, max_results=50):
    """
    Fetch user stories from Jira API
    """
    if not all([JIRA_CONFIG["base_url"], JIRA_CONFIG["email"], JIRA_CONFIG["api_token"]]):
        print("Jira configuration incomplete. Skipping Jira integration.")
        return []
    
    try:
        # Default JQL query for user stories
        if not jql_query:
            jql_query = f'project = {JIRA_CONFIG["project_key"]} AND issuetype = "User Story" AND status != Done ORDER BY priority DESC'
        
        url = f"{JIRA_CONFIG['base_url']}/rest/api/3/search"
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
        auth = (JIRA_CONFIG["email"], JIRA_CONFIG["api_token"])
        
        payload = {
            "jql": jql_query,
            "maxResults": max_results,
            "fields": ["summary", "description", "customfield_10014"]  # customfield_10014 is typically acceptance criteria
        }
        
        response = requests.post(url, headers=headers, auth=auth, json=payload)
        response.raise_for_status()
        
        data = response.json()
        stories = []
        
        for issue in data.get("issues", []):
            story_id = issue["key"]
            title = issue["fields"]["summary"]
            description = issue["fields"]["description"] or ""
            acceptance_criteria = issue["fields"].get("customfield_10014", "")
            
            # Combine description and acceptance criteria
            full_story = f"User Story: {title}\n\nDescription: {description}\n\nAcceptance Criteria:\n{acceptance_criteria}"
            
            stories.append({
                "id": story_id,
                "title": title,
                "story": full_story
            })
        
        print(f"✓ Retrieved {len(stories)} stories from Jira")
        return stories
        
    except Exception as e:
        print(f"Error fetching from Jira: {e}")
        return []

def generate_test_cases(user_story, story_id="TC"):
    """
    Send user story to AI API and generate risk-based test cases
    """
    # Determine the AI provider based on the configuration
    provider_name = AI_CONFIG["provider"]
    ai_provider = get_ai_provider(provider_name)
    
    # Add delay for rate limiting (especially for Gemini)
    if provider_name == "gemini":
        import time
        time.sleep(2)  # 2 second delay between requests

    return ai_provider.generate_test_cases(user_story, story_id)

def parse_markdown_table(markdown_text):
    """
    Parse the AI's markdown table output into a pandas DataFrame
    """
    try:
        # Find the table in the markdown text
        table_pattern = r'\|.*\|.*\|.*\|.*\|.*\|.*\|.*\|'
        table_lines = []
        
        for line in markdown_text.split('\n'):
            if re.match(table_pattern, line.strip()):
                table_lines.append(line.strip())
        
        if len(table_lines) < 3:  # Need header, separator, and at least one data row
            print("No valid table found in AI response")
            return None
        
        # Clean and parse the table
        cleaned_lines = []
        for line in table_lines:
            # Remove leading/trailing | and split by |
            cells = [cell.strip() for cell in line.strip('|').split('|')]
            if len(cells) == 7:  # Expected number of columns
                cleaned_lines.append(cells)
        
        if len(cleaned_lines) < 2:
            print("Invalid table structure")
            return None
        
        # Create DataFrame
        columns = cleaned_lines[0]  # Header row
        data = cleaned_lines[2:]    # Data rows (skip separator row)
        
        df = pd.DataFrame(data, columns=columns)
        
        # Clean up any markdown formatting
        for col in df.columns:
            df[col] = df[col].str.replace('**', '').str.replace('*', '').str.strip()
        
        return df
        
    except Exception as e:
        print(f"Error parsing markdown table: {e}")
        return None

def save_to_excel(all_test_cases, filename=None):
    """
    Save multiple DataFrames to Excel with proper formatting
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"test_cases_{timestamp}.xlsx"
    
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Add summary headers
        summary_headers = ["Story ID", "Story Title", "Total Test Cases", "High Risk", "Medium Risk", "Low Risk"]
        for col_num, header in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Process each story's test cases
        for story_data in all_test_cases:
            story_id = story_data["story_id"]
            story_title = story_data["story_title"]
            df = story_data["test_cases"]
            
            if df is None or len(df) == 0:
                continue
            
            # Create sheet for this story
            sheet_name = f"{story_id[:30]}"  # Excel sheet names limited to 31 chars
            ws = wb.create_sheet(sheet_name)
            
            # Write headers
            headers = list(df.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row_num, row in enumerate(df.values, 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add conditional formatting for risk levels
            high_risk_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            medium_risk_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            low_risk_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            
            risk_col = headers.index("Risk Level") + 1
            for row_num in range(2, len(df) + 2):
                risk_cell = ws.cell(row=row_num, column=risk_col)
                if risk_cell.value:
                    risk_level = str(risk_cell.value).upper()
                    if "HIGH" in risk_level:
                        risk_cell.fill = high_risk_fill
                    elif "MEDIUM" in risk_level:
                        risk_cell.fill = medium_risk_fill
                    elif "LOW" in risk_level:
                        risk_cell.fill = low_risk_fill
            
            # Add to summary sheet
            risk_counts = df['Risk Level'].value_counts()
            summary_row = [
                story_id,
                story_title,
                len(df),
                risk_counts.get('High', 0),
                risk_counts.get('Medium', 0),
                risk_counts.get('Low', 0)
            ]
            
            for col_num, value in enumerate(summary_row, 1):
                cell = summary_ws.cell(row=len(all_test_cases) + 2, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left")
        
        # Format summary sheet
        for column in summary_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            summary_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(filename)
        print(f"✓ Test cases saved to {filename}")
        return True
        
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False

def process_stories_bulk(stories, output_filename=None):
    """
    Process multiple stories in bulk
    """
    all_test_cases = []
    total_stories = len(stories)
    
    print(f"=== Processing {total_stories} stories ===")
    
    for i, story in enumerate(stories, 1):
        print(f"\n[{i}/{total_stories}] Processing: {story['title']}")
        
        # Generate test cases
        ai_response = generate_test_cases(story['story'], story['id'])
        
        if not ai_response:
            print(f"⚠️  Failed to generate test cases for {story['title']}")
            continue
        
        # Parse the response
        df = parse_markdown_table(ai_response)
        
        if df is None:
            print(f"⚠️  Failed to parse test cases for {story['title']}")
            continue
        
        all_test_cases.append({
            "story_id": story['id'],
            "story_title": story['title'],
            "test_cases": df
        })
        
        print(f"✓ Generated {len(df)} test cases")
        
        # Rate limiting to avoid API limits
        if i < total_stories:
            time.sleep(1)
    
    # Save all test cases to Excel
    if all_test_cases:
        success = save_to_excel(all_test_cases, output_filename)
        if success:
            print(f"\n=== Summary ===")
            total_cases = sum(len(tc['test_cases']) for tc in all_test_cases)
            print(f"Total stories processed: {len(all_test_cases)}")
            print(f"Total test cases generated: {total_cases}")
            
            # Overall risk distribution
            all_risk_levels = []
            for tc in all_test_cases:
                all_risk_levels.extend(tc['test_cases']['Risk Level'].tolist())
            
            risk_counts = pd.Series(all_risk_levels).value_counts()
            print(f"Risk distribution:")
            for risk, count in risk_counts.items():
                print(f"  {risk}: {count}")
    else:
        print("No test cases were generated successfully.")

def get_user_input_story():
    """
    Get user story and acceptance criteria interactively
    """
    print("\n=== Custom User Story Input ===")
    print("Enter your user story and acceptance criteria below.")
    print("Press Enter twice when done with each section.\n")
    
    # Get story ID
    story_id = input("Enter Story ID (e.g., US001): ").strip()
    if not story_id:
        story_id = "US001"
    
    # Get story title
    title = input("Enter Story Title: ").strip()
    if not title:
        title = "Custom User Story"
    
    # Get user story
    print("\nEnter your User Story (press Enter twice when done):")
    user_story_lines = []
    while True:
        line = input()
        if line == "" and user_story_lines and user_story_lines[-1] == "":
            break
        user_story_lines.append(line)
    
    user_story = "\n".join(user_story_lines[:-1])  # Remove the last empty line
    
    # Get acceptance criteria
    print("\nEnter your Acceptance Criteria (press Enter twice when done):")
    acceptance_lines = []
    while True:
        line = input()
        if line == "" and acceptance_lines and acceptance_lines[-1] == "":
            break
        acceptance_lines.append(line)
    
    acceptance_criteria = "\n".join(acceptance_lines[:-1])  # Remove the last empty line
    
    # Combine into full story
    full_story = f"""
User Story: {user_story}

Acceptance Criteria:
{acceptance_criteria}
"""
    
    return {
        "id": story_id,
        "title": title,
        "story": full_story
    }

def create_custom_story_file():
    """
    Create a template JSON file for custom stories
    """
    template = [
        {
            "id": "US001",
            "title": "Your Custom User Story",
            "story": """
User Story: As a [user type], I want to [action/feature] so that [benefit/value].

Acceptance Criteria:
1. [First acceptance criterion]
2. [Second acceptance criterion]
3. [Third acceptance criterion]
4. [Add more as needed]
"""
        }
    ]
    
    filename = "custom_stories.json"
    with open(filename, 'w') as f:
        json.dump(template, f, indent=2)
    
    print(f"✅ Created template file: {filename}")
    print("📝 Edit this file with your custom stories and run:")
    print(f"   python test_case_generator.py --mode bulk --stories {filename}")
    
    return filename

def main():
    """
    Main function with command line argument support
    """
    parser = argparse.ArgumentParser(description='Generate risk-based test cases from user stories')
    parser.add_argument('--mode', choices=['single', 'bulk', 'jira', 'interactive'], default='single',
                       help='Processing mode: single story, bulk processing, Jira integration, or interactive input')
    parser.add_argument('--jql', type=str, help='JQL query for Jira stories')
    parser.add_argument('--output', type=str, help='Output filename for Excel file')
    parser.add_argument('--stories', type=str, help='JSON file containing multiple stories')
    parser.add_argument('--story', type=str, help='Quick custom user story (use with --acceptance)')
    parser.add_argument('--acceptance', type=str, help='Quick custom acceptance criteria (use with --story)')
    
    args = parser.parse_args()
    
    print("=== Enhanced Test Case Generator ===")
    
    if args.mode == 'single':
        # Process single example story or custom story
        if args.story and args.acceptance:
            # Quick custom story input
            custom_story = {
                "id": "US001",
                "title": "Custom User Story",
                "story": f"""
User Story: {args.story}

Acceptance Criteria:
{args.acceptance}
"""
            }
            print("Processing custom story...")
            process_stories_bulk([custom_story], args.output)
        else:
            # Process single example story
            print("Processing single story...")
            process_stories_bulk([EXAMPLE_STORIES[0]], args.output)
        
    elif args.mode == 'bulk':
        # Process multiple stories
        if args.stories:
            try:
                with open(args.stories, 'r') as f:
                    stories = json.load(f)
                process_stories_bulk(stories, args.output)
            except Exception as e:
                print(f"Error loading stories from file: {e}")
        else:
            # Use example stories
            print("Processing example stories in bulk...")
            process_stories_bulk(EXAMPLE_STORIES, args.output)
            
    elif args.mode == 'jira':
        # Fetch and process stories from Jira
        print("Fetching stories from Jira...")
        jira_stories = get_jira_stories(args.jql)
        if jira_stories:
            process_stories_bulk(jira_stories, args.output)
        else:
            print("No stories retrieved from Jira. Check your configuration and JQL query.")

    elif args.mode == 'interactive':
        # Get user input for a single story
        print("=== Interactive User Story Input ===")
        user_story_data = get_user_input_story()
        print("\nGenerating test cases for:")
        print(f"  ID: {user_story_data['id']}")
        print(f"  Title: {user_story_data['title']}")
        print(f"  Story: {user_story_data['story']}")

        ai_response = generate_test_cases(user_story_data['story'], user_story_data['id'])

        if not ai_response:
            print("⚠️  Failed to generate test cases for the custom story.")
        else:
            df = parse_markdown_table(ai_response)
            if df is None:
                print("⚠️  Failed to parse test cases for the custom story.")
            else:
                print(f"\nGenerated {len(df)} test cases:")
                print(df.to_string(index=False))

                # Option to save to Excel
                save_prompt = input("\nWould you like to save these test cases to Excel? (y/n): ").strip().lower()
                if save_prompt == 'y':
                    save_to_excel([{"story_id": user_story_data['id'], "story_title": user_story_data['title'], "test_cases": df}], args.output)
                    print(f"✓ Test cases saved to {args.output}")
                else:
                    print("Test cases not saved.")

        # Option to create a custom story file
        create_prompt = input("\nWould you like to create a template file for custom stories? (y/n): ").strip().lower()
        if create_prompt == 'y':
            create_custom_story_file()

if __name__ == "__main__":
    main() 